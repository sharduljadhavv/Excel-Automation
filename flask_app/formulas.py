from datetime import datetime, timedelta
from .operations import get_original_format
from copy import copy
from openpyxl.styles import Alignment, PatternFill, Font

#AGEING
def apply_ageing(sheet, recon_date, start_row, end, column):
    try:
        for row in range(start_row, end):
            # Retrieve transaction date from column E for the current row
            date = sheet.cell(row=row, column=column).value

            # Check if the variable is a datetime object
            if date is not None and isinstance(date, datetime):
                # Calculate ageing
                ageing = abs((recon_date - date).days)
            else:
                # If date is not a datetime object, assume it's a string
                if date is not None and isinstance(date, str):
                    # Identify the original format of the date
                    original_format = get_original_format(date)
                    if original_format is not None:
                        # Convert the string to a datetime object using the identified format
                        date = original_format['datetime']
                    else:
                        # If the format cannot be identified, try a default format
                        date = datetime.strptime(date, '%d/%m/%Y %I:%M %p')

                # Calculate ageing
                ageing = abs((recon_date - date).days)

            sheet.cell(row=row, column=2).value = ageing
    except Exception as e:
        print(f"str{e}")

#TAT
def apply_TAT(sheet, start_row, end, column):
    try:
        for row in range(start_row, end):
            # Retrieve the value from column B for the current row
            value = sheet.cell(row=row, column=column).value

            # Apply the formula based on the value and update the cell in column A
            if value is not None:
                if value <= 3:
                    result = "0 to 3 Days"
                elif value <= 7:
                    result = "4 to 7 Days"
                elif value <= 15:
                    result = "8 to 15 Days"
                elif value <= 30:
                    result = "16 to 30 Days"
                elif value <= 60:
                    result = "31 to 60 Days"
                elif value <= 90:
                    result = "61 to 90 Days"
                elif value <= 180:
                    result = "91 to 180 Days"
                elif value <= 270:
                    result = "181 to 270 Days"
                elif value <= 360:
                    result = "271 to 360 Days"
                else:
                    result = "> 361 Days"

                sheet.cell(row=row, column=1).value = result
    except Exception as e:
        print(f"str{e}")

#SLA
def apply_SLA(sheet, start_row, end, column):
    try:
        for row in range(start_row, end):
            # Retrieve the value from column B for the current row
            value = sheet.cell(row=row, column=column).value

            # Apply the formula based on the value and update the cell in column C
            if value is not None:
                if value <= 4:
                    result = "Within SLA"
                else:
                    result = "Beyond SLA"   

                sheet.cell(row=row, column=3).value = result
    except Exception as e:
        print(f"str{e}")

#MONTHS
def apply_month(recon_date, sheet, start_row, end, column):
    try:
        for row in range(start_row, end):
            # Retrieve the value from column E for the current row
            value = sheet.cell(row=row, column=column).value

            if value is not None and not isinstance(value, str):
                # Calculate the date to compare
                compare_date = recon_date - \
                    timedelta(days=12*365/12)  # Subtract 12 months

                if compare_date - value > timedelta(days=0):
                    result = "< " + compare_date.strftime("%b_%Y")
                else:
                    result = value.strftime("%b_%Y")

            else:
                # If value is not a datetime object, assume it's a string
                if value is not None and isinstance(value, str):
                    # Identify the original format of the date
                    original_format = get_original_format(value)
                    if original_format is not None:
                        # Convert the string to a datetime object using the identified format
                        value = original_format['datetime']
                    else:
                        # If the format cannot be identified, try a default format
                        value = datetime.strptime(value, '%d/%m/%Y %I:%M %p')

                # Calculate the date to compare
                compare_date = recon_date - \
                    timedelta(days=365/12 * 12)  # Subtract 12 months

                if compare_date - value > timedelta(days=0):
                    result = "< " + compare_date.strftime("%b_%Y")
                else:
                    result = value.strftime("%b_%Y")


            # Update the cell in column D with the result
            sheet.cell(row=row, column=4).value = result
    except Exception as e:
        print(f"str{e}")

#VLOOKUPs
def knockoff_matching(sheet1, sheet2, sheet3, sheet4, destination_sheet, top_sheet_date):
    try:
        date_added = False
        # Format for the date
        date_format = '%m/%d/%Y'
        final_date = top_sheet_date.strftime(date_format)
    except Exception as e:
        print('str{e}')

    # =======================================================================#
    #ANNEX 1 and 4
    matched_data_1_4 = []
    try:
        amount_total_14 = 0
        start_row = destination_sheet.max_row + 1

        # Adding date in the first column of the row
        date_cell = destination_sheet.cell(row=start_row, column=1)
        date_cell.value = final_date
        date_cell.number_format = 'mm/dd/yyyy'
        date_cell.fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        date_cell.alignment = Alignment(horizontal='right')

        # Copy headers from Sheet 2
        for j in range(1, sheet1.max_column + 1):
            source_cell = sheet1.cell(row=5, column=j)
            destination_cell = destination_sheet.cell(row=start_row, column=j + 1)
            destination_cell.value = source_cell.value
            if source_cell.has_style:
                destination_cell.font = copy(source_cell.font)
                destination_cell.border = copy(source_cell.border)
                destination_cell.fill = copy(source_cell.fill)
                destination_cell.number_format = copy(source_cell.number_format)
                destination_cell.protection = copy(source_cell.protection)
                destination_cell.alignment = copy(source_cell.alignment)

        start_row += 1

        for i in range(6, sheet1.max_row + 1):  # Assuming headers are in row 5
            value_sheet1 = sheet1.cell(row=i, column=9).value

            if value_sheet1 not in ["N/A", None]:
                matching_value_found = False
                for row in sheet4.iter_rows(min_row=5, min_col=17, values_only=True):
                    if value_sheet1 == row[0]:
                        matching_value_found = True
                        matched_data_1_4.append((i, [sheet1.cell(row=i, column=j).value for j in range(1, sheet1.max_column + 1)]))
                        break

                if matching_value_found:
                    for j in range(1, sheet1.max_column + 1):
                        source_cell = sheet1.cell(row=i, column=j)
                        destination_cell = destination_sheet.cell(
                            row=start_row, column=j + 1)
                        destination_cell.value = source_cell.value
                        
                        if source_cell.has_style:
                            destination_cell.font = copy(source_cell.font)
                            destination_cell.border = copy(source_cell.border)
                            destination_cell.fill = copy(source_cell.fill)
                            destination_cell.number_format = copy(
                                source_cell.number_format)
                            destination_cell.protection = copy(
                                source_cell.protection)
                            destination_cell.alignment = copy(
                                source_cell.alignment)
                        
                      
                        if j == 9:
                            print(source_cell.value)
                            amount_total_14 += int(source_cell.value)

                        
                    start_row += 1
                    

        total_cell_14 = destination_sheet.cell(row=start_row, column=10)
        total_cell_14.value = amount_total_14
        total_cell_14.number_format = '0.00'
        total_cell_14.font = Font(bold=True)

        print("ANNEX 1 and 4 DONE")
    except Exception as e:
        print(f'{str(e)}')
        
    # ANNEX 4 and 1
    matched_data_4_1 = []
    try:
        amount_total_41 = 0
        start_row = destination_sheet.max_row + 1

        # Copy headers from Sheet 4
        for j in range(1, sheet4.max_column + 1):
            source_cell = sheet4.cell(row=5, column=j)
            destination_cell = destination_sheet.cell(
                row=start_row, column=j + 1)
            destination_cell.value = source_cell.value
            if source_cell.has_style:
                destination_cell.font = copy(source_cell.font)
                destination_cell.border = copy(source_cell.border)
                destination_cell.fill = copy(source_cell.fill)
                destination_cell.number_format = copy(source_cell.number_format)
                destination_cell.protection = copy(source_cell.protection)
                destination_cell.alignment = copy(source_cell.alignment)

        start_row += 1

        for i in range(6, sheet4.max_row + 1):  # Assuming headers are in row 5
            value_sheet4 = sheet4.cell(row=i, column=17).value

            if value_sheet4 not in ["N/A", None]:
                matching_value_found = False
                for row in sheet1.iter_rows(min_row=5, min_col=9, values_only=True):
                    if value_sheet4 == row[0]:
                        matching_value_found = True
                        matched_data_4_1.append((i, [sheet4.cell(row=i, column=j).value for j in range(1, sheet4.max_column + 1)]))
                        break

                if matching_value_found:
                    for j in range(1, sheet4.max_column + 1):
                        source_cell = sheet4.cell(row=i, column=j)
                        destination_cell = destination_sheet.cell(
                            row=start_row, column=j + 1)
                        destination_cell.value = source_cell.value
                        if source_cell.has_style:
                            destination_cell.font = copy(source_cell.font)
                            destination_cell.border = copy(source_cell.border)
                            destination_cell.fill = copy(source_cell.fill)
                            destination_cell.number_format = copy(
                                source_cell.number_format)
                            destination_cell.protection = copy(
                                source_cell.protection)
                            destination_cell.alignment = copy(
                                source_cell.alignment)

                        if j == 17:
                            print(source_cell.value)
                            amount_total_41 += int(source_cell.value)

                    start_row += 1

        total_cell_41 = destination_sheet.cell(row=start_row, column=18)
        total_cell_41.value = amount_total_41
        total_cell_41.number_format = '0.00'
        total_cell_41.font = Font(bold=True)

        print("ANNEX 4 and 1 DONE")
    except Exception as e:
        print(f'{str(e)}')
        
    # Remove matched rows from Sheet1
    for i in range(sheet1.max_row, 5, -1):
        for index, (row_index, data_row) in enumerate(matched_data_1_4):
            match_found = True
            for j, cell_value in enumerate(data_row, start=1):
                if sheet1.cell(row=i, column=j).value != cell_value:
                    match_found = False
                    break
            if match_found:
                sheet1.delete_rows(i)
                del matched_data_1_4[index]
                break

    # Remove matched rows from Sheet4
    for i in range(sheet4.max_row, 5, -1):
        for index, (row_index, data_row) in enumerate(matched_data_4_1):
            match_found = True
            for j, cell_value in enumerate(data_row, start=1):
                if sheet4.cell(row=i, column=j).value != cell_value:
                    match_found = False
                    break
            if match_found:
                sheet4.delete_rows(i)
                del matched_data_4_1[index]
                break
    # =======================================================================#
    
    # =======================================================================#
    # annex 2 and 3
    matched_data_2_3 = []
    try:
        amount_total_23 = 0
        start_row = destination_sheet.max_row + 1

        # Copy headers from Sheet 2
        for j in range(1, sheet2.max_column + 1):
            source_cell = sheet2.cell(row=5, column=j)
            destination_cell = destination_sheet.cell(row=start_row, column=j + 1)
            destination_cell.value = source_cell.value
            if source_cell.has_style:
                destination_cell.font = copy(source_cell.font)
                destination_cell.border = copy(source_cell.border)
                destination_cell.fill = copy(source_cell.fill)
                destination_cell.number_format = copy(source_cell.number_format)
                destination_cell.protection = copy(source_cell.protection)
                destination_cell.alignment = copy(source_cell.alignment)

        start_row += 1

        for i in range(6, sheet2.max_row + 1):  # Assuming headers are in row 5
            value_sheet2 = sheet2.cell(row=i, column=13).value

            if value_sheet2 not in ["N/A", None]:
                matching_value_found = False
                for row in sheet3.iter_rows(min_row=5, min_col=20, values_only=True):
                    if value_sheet2 == row[0]:
                        matching_value_found = True
                        matched_data_2_3.append((i, [sheet2.cell(row=i, column=j).value for j in range(1, sheet2.max_column + 1)]))
                        break

                if matching_value_found:
                    for j in range(1, sheet2.max_column + 1):
                        source_cell = sheet2.cell(row=i, column=j)
                        destination_cell = destination_sheet.cell(
                            row=start_row, column=j + 1)
                        destination_cell.value = source_cell.value
                        
                        if source_cell.has_style:
                            destination_cell.font = copy(source_cell.font)
                            destination_cell.border = copy(source_cell.border)
                            destination_cell.fill = copy(source_cell.fill)
                            destination_cell.number_format = copy(
                                source_cell.number_format)
                            destination_cell.protection = copy(
                                source_cell.protection)
                            destination_cell.alignment = copy(
                                source_cell.alignment)

                        if j == 17:
                            print(source_cell.value)
                            amount_total_23 += int(source_cell.value)
                        # if source_cell.value is not None:
                        #     print(source_cell.value)
                        #     try:
                        #         if isinstance(source_cell.value, (str, int, float)):
                        #             
                        #         else:
                        #             print("Non-numeric value found")
                        #     except ValueError:
                        #         print("Non-integer value found")

                
                    start_row += 1
                    

        total_cell_23 = destination_sheet.cell(row=start_row, column=18)
        total_cell_23.value = amount_total_23
        total_cell_23.number_format = '0.00'
        total_cell_23.font = Font(bold=True)

        print("ANNEX 2 and 3 DONE")
    except Exception as e:
        print(f'{str(e)}')

    # annex 3 and 2
    matched_data_3_2 = []
    try:
        amount_total_32 = 0
        start_row = destination_sheet.max_row + 1

        # Copy headers from Sheet 2
        for j in range(1, sheet3.max_column + 1):
            source_cell = sheet3.cell(row=5, column=j)
            destination_cell = destination_sheet.cell(row=start_row, column=j + 1)
            destination_cell.value = source_cell.value
            if source_cell.has_style:
                destination_cell.font = copy(source_cell.font)
                destination_cell.border = copy(source_cell.border)
                destination_cell.fill = copy(source_cell.fill)
                destination_cell.number_format = copy(source_cell.number_format)
                destination_cell.protection = copy(source_cell.protection)
                destination_cell.alignment = copy(source_cell.alignment)

        start_row += 1

        for i in range(6, sheet3.max_row + 1):  # Assuming headers are in row 5
            value_sheet3 = sheet3.cell(row=i, column=20).value

            if value_sheet3 not in ["N/A", None]:
                matching_value_found = False
                for row in sheet2.iter_rows(min_row=5, min_col=13, values_only=True):
                    if value_sheet3 == row[0]:
                        matching_value_found = True
                        matched_data_3_2.append((i, [sheet3.cell(row=i, column=j).value for j in range(1, sheet3.max_column + 1)]))
                        break

                if matching_value_found:
                    for j in range(1, sheet3.max_column + 1):
                        source_cell = sheet3.cell(row=i, column=j)
                        destination_cell = destination_sheet.cell(
                            row=start_row, column=j + 1)
                        destination_cell.value = source_cell.value
                        if source_cell.has_style:
                            destination_cell.font = copy(source_cell.font)
                            destination_cell.border = copy(source_cell.border)
                            destination_cell.fill = copy(source_cell.fill)
                            destination_cell.number_format = copy(
                                source_cell.number_format)
                            destination_cell.protection = copy(
                                source_cell.protection)
                            destination_cell.alignment = copy(
                                source_cell.alignment)

                        if j == 17:
                            print(source_cell.value)
                            amount_total_32 += int(source_cell.value)
                            
                    
                    start_row += 1
                    

        total_cell_32 = destination_sheet.cell(row=start_row, column=18)
        total_cell_32.value = amount_total_32
        total_cell_32.number_format = '0.00'
        total_cell_32.font = Font(bold=True)

        print("ANNEX 3 and 2 DONE")
    except Exception as e:
        print(f'{str(e)}')
    
    # Remove matched rows from Sheet1
    for i in range(sheet2.max_row, 5, -1):
        for index, (row_index, data_row) in enumerate(matched_data_2_3):
            match_found = True
            for j, cell_value in enumerate(data_row, start=1):
                if sheet2.cell(row=i, column=j).value != cell_value:
                    match_found = False
                    break
            if match_found:
                sheet2.delete_rows(i)
                del matched_data_2_3[index]
                break
            
    # Remove matched rows from Sheet4
    for i in range(sheet3.max_row, 5, -1):
        for index, (row_index, data_row) in enumerate(matched_data_3_2):
            match_found = True
            for j, cell_value in enumerate(data_row, start=1):
                if sheet3.cell(row=i, column=j).value != cell_value:
                    match_found = False
                    break
            if match_found:
                sheet3.delete_rows(i)
                del matched_data_3_2[index]
                break
    # =======================================================================#
    
    # =======================================================================# 
    # annex 3 and 4
    matched_data_3_4 = []
    try:
        amount_total_34 = 0
        start_row = destination_sheet.max_row + 1

        # Copy headers from Sheet 2
        for j in range(1, sheet3.max_column + 1):
            source_cell = sheet3.cell(row=5, column=j)
            destination_cell = destination_sheet.cell(
                row=start_row, column=j + 1)
            destination_cell.value = source_cell.value
            if source_cell.has_style:
                destination_cell.font = copy(source_cell.font)
                destination_cell.border = copy(source_cell.border)
                destination_cell.fill = copy(source_cell.fill)
                destination_cell.number_format = copy(
                    source_cell.number_format)
                destination_cell.protection = copy(source_cell.protection)
                destination_cell.alignment = copy(source_cell.alignment)

        start_row += 1

        for i in range(6, sheet3.max_row + 1):  # Assuming headers are in row 5
            value_sheet3 = sheet3.cell(row=i, column=20).value

            if value_sheet3 not in ["N/A", None]:
                matching_value_found = False
                for row in sheet4.iter_rows(min_row=5, min_col=20, values_only=True):
                    if value_sheet3 == row[0]:
                        matching_value_found = True
                        matched_data_3_4.append((i, [sheet3.cell(row=i, column=j).value for j in range(1, sheet3.max_column + 1)]))
                        break

                if matching_value_found:
                    for j in range(1, sheet3.max_column + 1):
                        source_cell = sheet3.cell(row=i, column=j)
                        destination_cell = destination_sheet.cell(
                            row=start_row, column=j + 1)
                        destination_cell.value = source_cell.value
                        if source_cell.has_style:
                            destination_cell.font = copy(source_cell.font)
                            destination_cell.border = copy(source_cell.border)
                            destination_cell.fill = copy(source_cell.fill)
                            destination_cell.number_format = copy(
                                source_cell.number_format)
                            destination_cell.protection = copy(
                                source_cell.protection)
                            destination_cell.alignment = copy(
                                source_cell.alignment)

                        if j == 17:
                            print(source_cell.value)
                            amount_total_34 += int(source_cell.value)

                    start_row += 1

        total_cell_34 = destination_sheet.cell(row=start_row, column=18)
        total_cell_34.value = amount_total_34
        total_cell_34.number_format = '0.00'
        total_cell_34.font = Font(bold=True)

        print("ANNEX 3 and 4 DONE")
    except Exception as e:
        print(f'{str(e)}')
        
    # annex 4 with 3
    matched_data_4_3 = []
    try:
        amount_total_43 = 0
        start_row = destination_sheet.max_row + 1

        # Copy headers from Sheet 2
        for j in range(1, sheet4.max_column + 1):
            source_cell = sheet4.cell(row=5, column=j)
            destination_cell = destination_sheet.cell(
                row=start_row, column=j + 1)
            destination_cell.value = source_cell.value
            if source_cell.has_style:
                destination_cell.font = copy(source_cell.font)
                destination_cell.border = copy(source_cell.border)
                destination_cell.fill = copy(source_cell.fill)
                destination_cell.number_format = copy(
                    source_cell.number_format)
                destination_cell.protection = copy(source_cell.protection)
                destination_cell.alignment = copy(source_cell.alignment)

        start_row += 1

        for i in range(6, sheet4.max_row + 1):  # Assuming headers are in row 5
            value_sheet4 = sheet4.cell(row=i, column=20).value

            if value_sheet4 not in ["N/A", None]:
                matching_value_found = False
                for row in sheet3.iter_rows(min_row=5, min_col=20, values_only=True):
                    if value_sheet4 == row[0]:
                        matching_value_found = True
                        matched_data_4_3.append((i, [sheet4.cell(row=i, column=j).value for j in range(1, sheet4.max_column + 1)]))
                        break

                if matching_value_found:
                    for j in range(1, sheet4.max_column + 1):
                        source_cell = sheet4.cell(row=i, column=j)
                        destination_cell = destination_sheet.cell(
                            row=start_row, column=j + 1)
                        destination_cell.value = source_cell.value
                        if source_cell.has_style:
                            destination_cell.font = copy(source_cell.font)
                            destination_cell.border = copy(source_cell.border)
                            destination_cell.fill = copy(source_cell.fill)
                            destination_cell.number_format = copy(
                                source_cell.number_format)
                            destination_cell.protection = copy(
                                source_cell.protection)
                            destination_cell.alignment = copy(
                                source_cell.alignment)

                        if j == 17:
                            print(source_cell.value)
                            amount_total_43 += int(source_cell.value)

                    start_row += 1

        total_cell_43 = destination_sheet.cell(row=start_row, column=18)
        total_cell_43.value = amount_total_43
        total_cell_43.number_format = '0.00'
        total_cell_43.font = Font(bold=True)

        print("ANNEX 4 and 3 DONE")
    except Exception as e:
        print(f'{str(e)}')
    
    # Remove matched rows from Sheet1
    for i in range(sheet3.max_row, 5, -1):
        for index, (row_index, data_row) in enumerate(matched_data_3_4):
            match_found = True
            for j, cell_value in enumerate(data_row, start=1):
                if sheet3.cell(row=i, column=j).value != cell_value:
                    match_found = False
                    break
            if match_found:
                sheet3.delete_rows(i)
                del matched_data_3_4[index]
                break
            
    # Remove matched rows from Sheet4
    for i in range(sheet4.max_row, 5, -1):
        for index, (row_index, data_row) in enumerate(matched_data_4_3):
            match_found = True
            for j, cell_value in enumerate(data_row, start=1):
                if sheet4.cell(row=i, column=j).value != cell_value:
                    match_found = False
                    break
            if match_found:
                sheet4.delete_rows(i)
                del matched_data_4_3[index]
                break
    # =======================================================================#
