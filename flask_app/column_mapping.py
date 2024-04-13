from openpyxl.utils import get_column_letter


#Column mapping for DR IN BANK Sheet
def create_mapping_dict_dr_bank(source_ws):
    try:
        column_names = ['value date', 'description', 'amount',
                        'reference no', 'branch name', 'c.d.falg']
        top_sheet_columns = [['E', 'J'], 'F', 'I', 'K', 'L', 'M']

        column_info = {}
        mapping_dict = {}

        for row in range(source_ws.min_row, source_ws.max_row + 1):
            for col in range(source_ws.min_column, source_ws.max_column + 1):
                cell_value = source_ws.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    normalized_value = cell_value.strip().lower()
                    if normalized_value in column_names and normalized_value not in column_info:
                        column_info[normalized_value] = {
                            'column': col, 'row': row + 1}

        # Create dictionary mapping
        for name, top_sheet_column in zip(column_names, top_sheet_columns):
            if name.lower() in column_info:
                column_index = column_info[name.lower()]['column']
                row_number = column_info[name.lower()]['row']
                cell = f"{get_column_letter(column_index)}{row_number}"
                mapping_dict[cell] = top_sheet_column

        return mapping_dict
    
    except Exception as e:
        print(f"str{e}")

#Column mapping for CR IN BANK Sheet
def create_mapping_dict_cr_bank(source_ws):
    try:
        column_names = ['external mid', 'external tid', 'upi merchant id', 'merchant name', 'merchant vpa', 'payer vpa', 'upi trxn id', 'order id',
                        'txn ref no. (rrn)', 'transaction req date', 'settlement date', 'currency', 'transaction amount', 'net amount', 'trans type', 'pay type', 'cr / dr', 'additional field 1', 'additional field 2', 'additional field 3', 'additional field 4', 'additional field 5', 'future free field 1']
        top_sheet_columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
                            'O', 'P', 'Q', 'W', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']

        column_info = {}
        mapping_dict = {}

        for row in range(source_ws.min_row, source_ws.max_row + 1):
            for col in range(source_ws.min_column, source_ws.max_column + 1):
                cell_value = source_ws.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    normalized_value = cell_value.strip().lower()
                    if normalized_value in column_names and normalized_value not in column_info:
                        column_info[normalized_value] = {
                            'column': col, 'row': row + 1}

        # Create dictionary mapping
        for name, top_sheet_column in zip(column_names, top_sheet_columns):
            if name.lower() in column_info:
                column_index = column_info[name.lower()]['column']
                row_number = column_info[name.lower()]['row']
                cell = f"{get_column_letter(column_index)}{row_number}"
                mapping_dict[cell] = top_sheet_column

        return mapping_dict
    
    except Exception as e:
        print(f"str{e}")

#Column mapping for DR IN LEDGER Sheet with ledgerwise input sheet
def create_mapping_dict_dr_ledger(source_ws):
    try:
        column_names = ['voucher date', 'voucherid', 'narration', 'reference', 'dealid', 'customer name', 'transaction', 'maker id',
                        'checker id', 'value_date', 'division', 'location', 'amount', 'dr/cr', 'loan agreement no', 'cheque no', 'cheque id', 'branch code', 'branch name', 'instalment no']

        top_sheet_columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
                            'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']

        column_info = {}
        mapping_dict = {}

        for row in range(source_ws.min_row, source_ws.max_row + 1):
            for col in range(source_ws.min_column, source_ws.max_column + 1):
                cell_value = source_ws.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    normalized_value = cell_value.strip().lower()
                    if normalized_value in column_names and normalized_value not in column_info:
                        column_info[normalized_value] = {
                            'column': col, 'row': row + 2}

        # Create dictionary mapping
        for name, top_sheet_column in zip(column_names, top_sheet_columns):
            if name.lower() in column_info:
                column_index = column_info[name.lower()]['column']
                row_number = column_info[name.lower()]['row']
                cell = f"{get_column_letter(column_index)}{row_number}"
                mapping_dict[cell] = top_sheet_column

        return mapping_dict
    
    except Exception as e:
        print(f"str{e}")

#Column mapping for DR IN LEDGER Sheet with pennant input sheet
def create_mapping_dict_dr_ledger_pennant(source_ws):
    try:
        column_names = ['voucherdate', 'voucherid', 'systemname', 'customername', 'makerid',
                        'checkerid', 'valuedate', 'dramt', 'agreementno', 'cheque_no', 'chequeid', 'branchid', 'branch_name']

        top_sheet_columns = ['E', 'F', 'G', 'J', 'L',
                            'M', 'N', 'Q', 'S', 'T', 'U', 'V', 'W']

        column_info = {}
        mapping_dict = {}

        for row in range(source_ws.min_row, source_ws.max_row + 1):
            for col in range(source_ws.min_column, source_ws.max_column + 1):
                cell_value = source_ws.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    normalized_value = cell_value.strip().lower()
                    if normalized_value in column_names and normalized_value not in column_info:
                        column_info[normalized_value] = {
                            'column': col, 'row': row + 2}

        # Create dictionary mapping
        for name, top_sheet_column in zip(column_names, top_sheet_columns):
            if name.lower() in column_info:
                column_index = column_info[name.lower()]['column']
                row_number = column_info[name.lower()]['row']
                cell = f"{get_column_letter(column_index)}{row_number}"
                mapping_dict[cell] = top_sheet_column

        return mapping_dict
    
    except Exception as e:
        print(f"str{e}")

#Column mapping for CR IN LEDGER Sheet with ledgerwise input sheet
def create_mapping_dict_cr_ledger(source_ws):
    try:
        column_names = ['voucher date', 'voucherid', 'narration', 'reference', 'dealid', 'customer name', 'transaction', 'maker id',
                        'checker id', 'value_date', 'division', 'location', 'amount', 'dr/cr', 'loan agreement no', 'cheque no', 'cheque id', 'branch code', 'branch name', 'instalment no']

        top_sheet_columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
                            'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']

        column_info = {}
        mapping_dict = {}

        for row in range(source_ws.min_row, source_ws.max_row + 1):
            for col in range(source_ws.min_column, source_ws.max_column + 1):
                cell_value = source_ws.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    normalized_value = cell_value.strip().lower()
                    if normalized_value in column_names and normalized_value not in column_info:
                        column_info[normalized_value] = {
                            'column': col, 'row': row + 2}

        # Create dictionary mapping
        for name, top_sheet_column in zip(column_names, top_sheet_columns):
            if name.lower() in column_info:
                column_index = column_info[name.lower()]['column']
                row_number = column_info[name.lower()]['row']
                cell = f"{get_column_letter(column_index)}{row_number}"
                mapping_dict[cell] = top_sheet_column

        return mapping_dict
    
    except Exception as e:
        print(f"str{e}")

# Column mapping for CR IN LEDGER Sheet with Pennant input sheet
def create_mapping_dict_cr_ledger_pennant(source_ws):
    try:
        column_names = ['voucherdate', 'voucherid', 'systemname', 'customername', 'makerid',
                        'checkerid', 'valuedate', 'cramt', 'agreementno', 'cheque_no', 'chequeid', 'branchid', 'branch_name']

        top_sheet_columns = ['E', 'F', 'G', 'J', 'L',
                            'M', 'N', 'Q', 'S', 'T', 'U', 'V', 'W']

        column_info = {}
        mapping_dict = {}

        for row in range(source_ws.min_row, source_ws.max_row + 1):
            for col in range(source_ws.min_column, source_ws.max_column + 1):
                cell_value = source_ws.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    normalized_value = cell_value.strip().lower()
                    if normalized_value in column_names and normalized_value not in column_info:
                        column_info[normalized_value] = {
                            'column': col, 'row': row + 2}

        # Create dictionary mapping
        for name, top_sheet_column in zip(column_names, top_sheet_columns):
            if name.lower() in column_info:
                column_index = column_info[name.lower()]['column']
                row_number = column_info[name.lower()]['row']
                cell = f"{get_column_letter(column_index)}{row_number}"
                mapping_dict[cell] = top_sheet_column

        return mapping_dict
    
    except Exception as e:
        print(f"str{e}")
