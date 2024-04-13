from openpyxl.styles import Alignment
from datetime import datetime
import pandas as pd

#To get the start and end row number that contains data in input sheet
def start_end_row(source_ws, term, from_line=1, column=1):
    try:
        input_start_row = None
        input_end_row = None
        blank_row_count = 0

        for row in range(source_ws.min_row, source_ws.max_row + 1):
            cell_value = source_ws.cell(row=row, column=column).value
            if isinstance(cell_value, str) and cell_value.lower() == term:
                input_start_row = row + from_line  # The next row after the one containing "term"
                blank_row_count = 0  # Reset the blank row count when starting row is found
            elif input_start_row is not None:
                if cell_value is not None:
                    input_end_row = row  # Update the end row if a non-blank row is found
                    blank_row_count = 0  # Reset the blank row count when a non-blank row is found
                else:
                    blank_row_count += 1  # Increment blank row count if the current row is blank
                    if blank_row_count >= 2:  # Check if more than two consecutive blank rows are encountered
                        break

        return {'start': input_start_row, 'end': input_end_row}
    
    except Exception as e:
        print(f"str{e}")

#To align content in the sheet
def align_to_center(sheet, horizontal='center', vertical='center'):
    try:
        # Iterate over all columns and rows to set alignment to center
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
    except Exception as e:
        print(f"Error occurred: {e}")

#To change the data type to integer
def to_integer(start_row, end,  sheet, column_letter):
    try:
        # Iterate over rows starting from the 6th row in the specified column and convert values to integers
        for row in range(start_row, end):
            cell = sheet[column_letter + str(row)]
            if cell.value is not None:  # Check if the cell has a value
                try:
                    cell.value = pd.to_numeric(cell.value, errors='coerce')
                except ValueError:
                    print(
                        f"Unable to convert value {cell.value} to an integer in row {row}.")
            else:
                print("cell value is None")
    except Exception as e:
        print(f"str{e}")

#To change the data type to string
def to_string(start_row, end_row, sheet, column_letter):
    try:
        # Iterate over rows starting from the specified start_row and ending at end_row
        for row in range(start_row, end_row + 1):  # Add 1 to include the end_row
            cell = sheet[column_letter + str(row)]
            if cell.value is not None:  # Check if the cell has a value
                try:
                    # Try to convert the value to a string
                    cell.value = str(cell.value)
                except Exception as e:
                    # If conversion to string fails, handle the error here
                    print(
                        f"Unable to convert value {cell.value} to a string in row {row}.")
            else:
                print("cell value is None")
    except Exception as e:
        print(f"str{e}")

#To convert the datetime format. e.g. 24-Feb-24
def format_dates_in_sheet(start_row, end_row, sheet, column_one, column_two):
    try:
        date_format = '%d-%b-%y'
        for row in range(start_row, end_row + 1):
            # Retrieve the value from the cell
            cell_value1 = sheet[f'{column_one}{row}'].value
            cell_value2 = sheet[f'{column_two}{row}'].value

            # Check if the cell value is not empty and is a date
            if cell_value1 is not None and isinstance(cell_value1, datetime):
                # Convert the date to the desired format and assign it back to the cell
                formatted_date1 = cell_value1.strftime(date_format)
                sheet[f'{column_one}{row}'] = formatted_date1
            elif isinstance(cell_value1, str):
                # Identify the original format of the date
                original_format1 = get_original_format(cell_value1)
                if original_format1 is not None:
                    # Convert the string to datetime object
                    datetime_obj1 = original_format1['datetime']
                    # Format the datetime object to the desired format
                    formatted_date1 = datetime_obj1.strftime(date_format)
                    sheet[f'{column_one}{row}'] = formatted_date1

            if cell_value2 is not None and isinstance(cell_value2, datetime):
                # Convert the date to the desired format and assign it back to the cell
                formatted_date2 = cell_value2.strftime(date_format)
                sheet[f'{column_two}{row}'] = formatted_date2
            elif isinstance(cell_value2, str):
                # Identify the original format of the date
                original_format2 = get_original_format(cell_value2)
                if original_format2 is not None:
                    # Convert the string to datetime object
                    datetime_obj2 = original_format2['datetime']
                    # Format the datetime object to the desired format
                    formatted_date2 = datetime_obj2.strftime(date_format)
                    sheet[f'{column_two}{row}'] = formatted_date2
    except Exception as e:
        print(f"str{e}")

#To add "DR" to the column
def add_dr_to_column(sheet, start_row, end, column_name):
    try:
        for row in range(start_row, end):
            sheet[f'{column_name}{row}'] = 'DR'
    except Exception as e:
        print(f"str{e}")

#To add "CR" to the column
def add_cr_to_column(sheet, start_row, end, column_name):
    try:
        for row in range(start_row, end):
            e_column_value = sheet[f'E{row}'].value
            if e_column_value is not None and e_column_value != '':
                sheet[f'{column_name}{row}'] = 'CR'
    except Exception as e:
        print(f"str{e}")

def add_remarks_reason_fpr(sheet, start_row, end):
    try:
        for row in range(start_row, end + 1):
            e_column_value = sheet[f'E{row}'].value
            if e_column_value is not None and e_column_value != '' and row <= sheet.max_row: 
                sheet[f'Y{row}'] = 'credit not received'
                sheet[f'Z{row}'] = 'credit not received'
                sheet[f'AA{row}'] = 'service provider'
            else:
                print(f"Row {row} exceeds the maximum row count in the sheet.")
    except Exception as e:
        print(f"str{e}")

def find_new_row(sheet):
    try:
        last_row = 0
        for row in range(sheet.max_row, 0, -1):
            if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
                last_row = row
                break
        #New row after last 
        new_row = last_row + 1
        return new_row
    except Exception as e:
        print(f"str{e}")

#for advance use
def get_original_format(datetime_str):
    try:
        possible_formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%d',
            '%d-%m-%Y %H:%M:%S',
            '%d-%m-%Y %H:%M',
            '%d-%m-%Y',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y %H:%M',
            '%m/%d/%Y',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y',
            '%d/%m/%Y %I:%M %p',  
            '%d-%b-%Y %H:%M:%S',  
            '%d-%b-%Y',           
            '%m/%d/%Y',           
        ]

        for fmt in possible_formats:
            try:
                datetime_obj = datetime.strptime(datetime_str, fmt)
                return {'format': fmt, 'datetime': datetime_obj}
            except ValueError:
                pass

        return None
    except Exception as e:
        print(f"str{e}")