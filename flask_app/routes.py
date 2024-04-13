from flask import render_template, url_for, redirect, flash, request, send_file, jsonify
from flask_app import app, db, bcrypt
from .models import ProcessedFiles, User
import os, secrets
from .forms import LoginForm, RegistrationForm
from openpyxl import load_workbook
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl.utils import get_column_letter
import time
#different modules to perform repeated task
from .column_mapping import *
from .operations import *
from .formulas import *
from .load_workbook import *
import re
from flask_login import login_user, current_user, logout_user, login_required
#=================================imports=============================#

# Error handler for 404
@app.errorhandler(404)
def error_404(error):
    return render_template('error_pages/404.html'), 404

# Error handler for 404
@app.errorhandler(500)
def error_500(error):
    return render_template('error_pages/500.html'), 500

#================================Error pages============================#
@app.route("/login", methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and bcrypt.check_password_hash(user.password, form.password.data):
            login_user(user)
            flash('You have been logged in successfully!', 'success')
            return redirect('home')
        else:
            flash('Oops! It seems there was an error with your login. Please ensure your username and password are correct.', 'danger')
    return render_template('authentication/login.html', title='HDFC 4301 | Login', form=form)

@app.route("/register", methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = RegistrationForm()
    if request.method == 'POST':
        if form.validate_on_submit():
            hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
            user = User(username=form.username.data, email=form.email.data, password=hashed_password)
            db.session.add(user)
            db.session.commit()
            flash(f'Hey {form.username.data}, Your Account has been created.', 'success')
            return redirect(url_for('login'))
        else:
            flash(f'Oops! It seems there was an error with your registration. Please fill everything correctly', 'danger')
    return render_template('authentication/register.html', title='HDFC 4301 | Register', form=form)

@app.route("/logout")
def logout():
    logout_user()
    flash("You've been successfully logged out from your account!", 'success')
    return redirect(url_for('login'))
#=================================Auth routes==========================#

@app.route("/", methods=['GET'])
@app.route("/home", methods=['GET'])
def home():
    try:
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        files = ProcessedFiles.query.order_by(ProcessedFiles.created_at)
        return render_template('home.html', files=files, title="HDFC 4310 | Dashboard")
    except Exception as e:
        print(f"str{e}")

@app.route("/get_data/", methods=['POST'])
def get_data():
    start = time.time()
    try:
        #request.files.get to get a file.
        #we're targetting "name" attribute in the html form to get the file from the frontend.
        reconed_file = request.files.get('reconed-file')
        tejashree_file = request.files.get('bank-file')
        merchant_file = request.files.get('merchant-file')
        ledgerwise_file = request.files.get('ledger-file')
        pennant_file = request.files.get('pennant-file')
        #request.form.get to get date
        date_string = request.form.get('recon_date')
    except Exception as e:
        print(f"str{e}")

    #==================================main processing begin=============================#
    if reconed_file and tejashree_file and merchant_file and ledgerwise_file and pennant_file and date_string:
        
        # Convert the datetime string to a datetime object
        recon_date = datetime.strptime(date_string, '%Y-%m-%d')
        
        #read reconed file and build file path to the destination directory
        try:
            start = time.time()
            load_data = load_mastersheet(file=reconed_file)
            end = time.time()
            total = end - start
            print(total)
            file_path = os.path.join(app.root_path, 'media', 'processed files', f'BRS -{recon_date.year}-{recon_date.month}-{recon_date.day}- (HDFCCOLL4310) -BFL-HDFC-UPI COLLECTION INTEGRATION 4310.xlsx')
        except Exception as e:
            print(f"str{e}")
        
        try:
            #DR IN BANK
            if tejashree_file:
                #load mastersheet and input file
                source_ws = load_input_file(tejashree_file)
                workbook = load_data.get('workbook')
                
                #adding input date to the top sheet
                workbook['TOP SHEET']['D2'].value = recon_date
                
                print('NEW DATE ADDED TO THE TOP SHEET')
                
                sheet = workbook['DR IN BANK']

                #get mapping dictionary of source and target columns
                mapped_dictionary = create_mapping_dict_dr_bank(source_ws)

                #source file
                start_end = start_end_row(source_ws, 'transaction date')

                #target file row
                start_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
                        start_row = row + 1
                        break
                
                print("start row for DR IN BANK is", start_row)

                #append data to the target columns    
                for source_cell, target_columns in mapped_dictionary.items():
                    source_column = source_ws[source_cell].column
                    target_row = start_row  # Reset target row index for each column

                    # Iterate through rows from start_row to end_row in the source worksheet
                    for source_row in range(start_end.get('start'), start_end.get('end') + 1):
                        d_column = None  # Initialize d_column variable to None

                        # Iterate through cells in the current row to find the column containing "D"
                        for col in range(1, source_ws.max_column + 1):
                            cell_value = source_ws.cell(row=source_row, column=col).value
                            if cell_value == "D":
                                d_column = col
                                break  # Exit loop once "D" is found

                        if d_column is not None:
                            # Get the value from the source cell in the dynamically determined column
                            source_value = source_ws.cell(
                                row=source_row, column=source_column).value

                            # Loop over each target column for the source cell
                            for target_column in target_columns:
                                target_cell = f'{target_column}{target_row}'
                                sheet[target_cell].value = source_value
                                # print(f"Copied {source_value} to {target_cell}")
                            target_row += 1  # Increment target row index

                #to add 'DR' to a column    
                add_dr_to_column(sheet=sheet, start_row=start_row, end=sheet.max_row + 1, column_name='M')

                #Apply ageing
                apply_ageing(sheet, recon_date, start_row, sheet.max_row + 1, column=5)

                #Apply TAT
                apply_TAT(sheet, start_row, sheet.max_row + 1, column=2)

                #Apply SLA
                apply_SLA(sheet, start_row, sheet.max_row + 1, column=2)

                #MONTHS
                apply_month(recon_date, sheet, start_row, sheet.max_row + 1, column=5)

                # to change the datetime format
                format_dates_in_sheet(start_row, sheet.max_row + 1, sheet, column_one='E', column_two='J')

                # alignment to center
                align_to_center(sheet, horizontal='center', vertical='center')
                    
                for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row + 1, min_col=6, max_col=6, values_only=True):
                    if row[0] is not None:
                        cell_value = str(row[0])

                        # Define regular expression pattern to extract only matched text
                        pattern = re.compile(r'\b[^\W\d_]+\b')

                        # Find all matched text in the cell value
                        matched_text = pattern.findall(cell_value)

                        if matched_text:
                            # Join the matched text into a single string
                            matched_text = ' '.join(matched_text)

                            # Set the matched text to column 14
                            sheet.cell(row=start_row, column=14, value=matched_text)
                            sheet.cell(row=start_row, column=15, value=matched_text)
                            sheet.cell(row=start_row, column=16, value="Recon Team")
                        else:
                            print('No match found for row:', row[0].row)
                    start_row += 1  # Increment row counter if needed

                workbook.save(file_path)

                print('DR IN BANK - DONE')
            
            #CR IN BANK
            if merchant_file:  
                # load mastersheet and input file
                source_ws = load_input_file(merchant_file)
                workbook = load_data.get('workbook')
                sheet = workbook['CR IN BANK']

                #get mapping dictionary of source and target columns
                mapped_dictionary = create_mapping_dict_cr_bank(source_ws)

                #source file
                start_end = start_end_row(source_ws, 'external mid')

                #target file row
                start_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
                        start_row = row + 1
                        break
                
                print("start row for CR IN BANK is", start_row)

                # Iterate over the cell mappings
                for source_cell, target_column in mapped_dictionary.items():
                    source_column = source_ws[source_cell].column
                    target_row = start_row  # Start from the specified start row
                    for source_cell in source_ws.iter_rows(min_row=start_end.get('start'), max_row=start_end.get('end'), min_col=source_column, max_col=source_column):
                        target_cell = f'{target_column}{target_row}'
                        sheet[target_cell].value = source_cell[0].value
                        target_row += 1  # Increment target row index

                # Apply ageing
                apply_ageing(sheet, recon_date, start_row, sheet.max_row + 1, column=15)

                # Apply TAT
                apply_TAT(sheet, start_row, sheet.max_row + 1, column=2)

                # Apply SLA
                apply_SLA(sheet, start_row, sheet.max_row + 1, column=2)

                # MONTHS
                apply_month(recon_date, sheet, start_row, sheet.max_row + 1, column=15)

                #change the datatype to integer
                to_integer(start_row, sheet.max_row + 1, sheet, "Q")

                #change the datatype to integer
                to_integer(start_row, sheet.max_row + 1, sheet, "W")

                # Specify the source and target date formats
                format_dates_in_sheet(start_row, sheet.max_row + 1, sheet, column_one='N', column_two='O')

                #alignment to center
                align_to_center(sheet, horizontal='center', vertical='center')
                
                last_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
                        last_row = row
                        break

                for row in range(start_row, last_row + 1):
                    if row <= sheet.max_row:
                        sheet[f'AH{row}'] = 'Receipt not updated'
                        sheet[f'AI{row}'] = 'Receipt not updated'
                        sheet[f'AJ{row}'] = 'LAS Team/BFL Transaction Team'
                    else:
                        print(f"Row {row} exceeds the maximum row count in the sheet.")

                workbook.save(file_path)

                print('CR IN BANK - DONE')

            #DR IN LEDGER using Ledgerwise and pennant file
            if ledgerwise_file and pennant_file:
                # load mastersheet and input file
                source_ws = load_input_file(ledgerwise_file)
                pennant_source_ws = load_input_file(pennant_file)
                workbook = load_data.get('workbook')
                sheet = workbook['DR IN LEDGER']

                # get mapping dictionary of source and target columns
                mapped_dictionary = create_mapping_dict_dr_ledger(source_ws)

                # source file
                start_end = start_end_row(source_ws, 'voucher date', from_line=2)

                #target file row for ledgerwise data
                start_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
                        start_row = row + 1
                        break
                
                print("start row for DR IN LEDGER with LEDGERWISE file", start_row)

                for source_cell, target_columns in mapped_dictionary.items():
                    source_column = source_ws[source_cell].column
                    target_row = start_row  # Reset target row index for each column

                    # Iterate through rows from start_row to end_row in the source worksheet
                    for source_row in range(start_end.get('start'), start_end.get('end') + 1):
                        d_column = None  # Initialize d_column variable to None

                        # Iterate through cells in the current row to find the column containing "D"
                        for col in range(1, source_ws.max_column + 1):
                            cell_value = source_ws.cell(row=source_row, column=col).value
                            if cell_value == "DR":
                                d_column = col
                                break  # Exit loop once "D" is found

                        if d_column is not None:
                            # Get the value from the source cell in the dynamically determined column
                            source_value = source_ws.cell(
                                row=source_row, column=source_column).value

                            # Loop over each target column for the source cell
                            for target_column in target_columns:
                                target_cell = f'{target_column}{target_row}'
                                sheet[target_cell].value = source_value
                                # print(f"Copied {source_value} to {target_cell}")
                            target_row += 1  # Increment target row index

                # Add dr to the column
                add_dr_to_column(sheet=sheet, start_row=start_row, end=sheet.max_row + 1, column_name='R')
                
                # Apply ageing
                apply_ageing(sheet, recon_date, start_row, sheet.max_row + 1, column=5)

                # Apply TAT
                apply_TAT(sheet, start_row, sheet.max_row + 1, column=2)

                # Apply SLA
                apply_SLA(sheet, start_row, sheet.max_row + 1, column=2)

                # MONTHS
                apply_month(recon_date, sheet, start_row, sheet.max_row + 1, column=5)

                # to string
                to_string(start_row, sheet.max_row, sheet, 'T')

                #to change the datetime format
                format_dates_in_sheet(start_row, sheet.max_row + 1, sheet, column_one='E', column_two='N')

                workbook.save(file_path)

                print('DR IN LEDGER (LEDGERWISE FILE) - DONE')

                #======================================pennant======================================#

                # target file row for pennant data
                pennant_start_row = find_new_row(sheet)

                # get mapping dictionary of source and target columns
                mapped_dictionary_pennant = create_mapping_dict_dr_ledger_pennant(pennant_source_ws)

                # start and end data row from source file
                pennant_start_end = start_end_row(pennant_source_ws, term='voucherdate', from_line=2, column=2)

                # Iterate over the cell mappings
                for source_cell, target_column in mapped_dictionary_pennant.items():
                    source_column = pennant_source_ws[source_cell].column
                    target_row = pennant_start_row  # Start from the specified start row
                    for source_cell in pennant_source_ws.iter_rows(min_row=pennant_start_end.get('start'), max_row=pennant_start_end.get('end'), min_col=source_column, max_col=source_column):
                        target_cell = f'{target_column}{target_row}'
                        sheet[target_cell].value = source_cell[0].value
                        target_row += 1  # Increment target row index

                add_dr_to_column(sheet=sheet, start_row=pennant_start_row, end=sheet.max_row + 1, column_name='R')

                # Apply ageing
                apply_ageing(sheet, recon_date, pennant_start_row, sheet.max_row + 1, column=5)

                # Apply TAT
                apply_TAT(sheet, pennant_start_row, sheet.max_row + 1, column=2)

                # Apply SLA
                apply_SLA(sheet, pennant_start_row, sheet.max_row + 1, column=2)

                # MONTHS
                apply_month(recon_date, sheet, pennant_start_row, sheet.max_row + 1, column=5)

                # to string
                to_string(pennant_start_row, sheet.max_row, sheet, 'T')

                # to change the datetime format
                format_dates_in_sheet(pennant_start_row, sheet.max_row + 1, sheet, column_one='E', column_two='N')

                # alignment to center
                align_to_center(sheet, horizontal='center', vertical='center')
                
                # add remarks, reason, fpr
                last_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
                        last_row = row
                        break
                
                add_remarks_reason_fpr(sheet, 6, last_row)

                workbook.save(file_path)

                print('DR IN LEDGER (PENNANT FILE) - DONE')
            
            #CR IN LEDGER using Ledgerwise and pennant file
            if ledgerwise_file and pennant_file:
                # load mastersheet and input file
                source_ws = load_input_file(ledgerwise_file)
                pennant_source_ws = load_input_file(pennant_file)
                workbook = load_data.get('workbook')
                sheet = workbook['CR IN LEDGER']

                # get mapping dictionary of source and target columns
                mapped_dictionary = create_mapping_dict_cr_ledger(source_ws)

                # source file
                start_end = start_end_row(source_ws, 'voucher date', from_line=2)

                # target file row for ledgerwise data
                start_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
                        start_row = row + 1
                        break
                
                print("start row for CR IN LEDGER (LEDGERWISE file)", start_row)

                for source_cell, target_columns in mapped_dictionary.items():
                    source_column = source_ws[source_cell].column
                    target_row = start_row  # Reset target row index for each column

                    # Iterate through rows from start_row to end_row in the source worksheet
                    for source_row in range(start_end.get('start'), start_end.get('end') + 1):
                        d_column = None  # Initialize d_column variable to None

                        # Iterate through cells in the current row to find the column containing "D"
                        for col in range(1, source_ws.max_column + 1):
                            cell_value = source_ws.cell(row=source_row, column=col).value
                            if cell_value == "CR":
                                d_column = col
                                break  # Exit loop once "D" is found

                        if d_column is not None:
                            # Get the value from the source cell in the dynamically determined column
                            source_value = source_ws.cell(
                                row=source_row, column=source_column).value

                            # Loop over each target column for the source cell
                            for target_column in target_columns:
                                target_cell = f'{target_column}{target_row}'
                                sheet[target_cell].value = source_value
                                # print(f"Copied {source_value} to {target_cell}")
                            target_row += 1  # Increment target row index

                # add cr to the column
                add_cr_to_column(sheet=sheet, start_row=start_row, end=sheet.max_row + 1, column_name='R')

                # Apply ageing
                apply_ageing(sheet, recon_date, start_row, sheet.max_row + 1, column=5)

                # Apply TAT
                apply_TAT(sheet, start_row, sheet.max_row + 1, column=2)

                # Apply SLA
                apply_SLA(sheet, start_row, sheet.max_row + 1, column=2)

                # MONTHS
                apply_month(recon_date, sheet, start_row, sheet.max_row + 1, column=5)

                # to string
                to_string(start_row, sheet.max_row, sheet, 'T')

                # to change the datetime format
                format_dates_in_sheet(start_row, sheet.max_row + 1, sheet, column_one='E', column_two='N')

                # for closing balance in ledgewise
                for row in source_ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            if cell.value.lower() == 'closing balance':
                                # print(cell.row, cell.column, cell.value)
                                if cell.row + 1 is not None:
                                    closing_balance = source_ws.cell(row=cell.row + 1, column=cell.column).value
                                    int_closing_balance = pd.to_numeric(closing_balance, errors='coerce')
                                    negated_balance = -int_closing_balance if not pd.isna(int_closing_balance) else None
                                    workbook['TOP SHEET']['E8'].value = negated_balance

                workbook.save(file_path)

                print('CR IN LEDGER (LEDGERWISE FILE) - DONE')

                #========================================pennant=============================#

                # target file row for pennant data
                pennant_start_row = find_new_row(sheet)

                #get mapped dictionary of source and target sheet columns
                mapped_dictionary_pennant = create_mapping_dict_cr_ledger_pennant(pennant_source_ws)

                # source file
                pennant_start_end = start_end_row(pennant_source_ws, term='voucherdate', from_line=2, column=2)

                # Iterate over the cell mappings
                for source_cell, target_column in mapped_dictionary_pennant.items():
                    source_column = pennant_source_ws[source_cell].column
                    target_row = pennant_start_row  # Start from the specified start row
                    
                    # Iterate through rows from start_row to end_row in the source worksheet
                    for source_row in range(start_end.get('start'), start_end.get('end') + 1):
                        # Check if the value in the "cramt header" column is not zero
                        if source_ws.cell(row=source_row, column=source_column).value != 0:
                            # Loop over each target column for the source cell
                            for target_column in target_columns:
                                # Copy entire row to the target sheet
                                for col in range(1, source_ws.max_column + 1):
                                    source_value = source_ws.cell(row=source_row, column=col).value
                                    target_cell = f'{target_column}{target_row}'
                                    sheet[target_cell].value = source_value
                                target_row += 1  # Increment target row index

                add_cr_to_column(sheet=sheet, start_row=pennant_start_row, end=sheet.max_row + 1, column_name='R')

                # Apply ageing
                apply_ageing(sheet, recon_date, pennant_start_row, sheet.max_row + 1, column=5)

                # Apply TAT
                apply_TAT(sheet, pennant_start_row, sheet.max_row + 1, column=2)

                # Apply SLA
                apply_SLA(sheet, pennant_start_row, sheet.max_row + 1, column=2)

                # MONTHS
                apply_month(recon_date, sheet, pennant_start_row, sheet.max_row + 1, column=5)

                # to string
                to_string(pennant_start_row, sheet.max_row, sheet, 'T')

                # to change the datetime format
                format_dates_in_sheet(pennant_start_row, sheet.max_row + 1, sheet, column_one='E', column_two='N')

                # alignment to center
                align_to_center(sheet, horizontal='center', vertical='center')

                # closing balance to pennant = closing penn dr blnc - closing penn cr blnc
                dramt = None
                cramt = None

                for row in pennant_source_ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            if cell.value.lower() == 'dramt':
                                column_letter = get_column_letter(cell.column)
                                dramt = pennant_source_ws[f"{column_letter}{pennant_source_ws.max_row}"].value

                            elif cell.value.lower() == 'cramt':
                                column_letter = get_column_letter(cell.column)
                                cramt = pennant_source_ws[f"{column_letter}{pennant_source_ws.max_row}"].value

                total = dramt - cramt

                # Assign value to E9 cell in top sheet
                workbook['TOP SHEET']['E9'].value = total
                
                # add remarks, reason, fpr
                last_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
                        last_row = row
                        break
                add_remarks_reason_fpr(sheet, 6, last_row)
                
                # Vlookup and Knockoffs
                knockoff_matching(sheet1=workbook['DR IN BANK'],
                                  sheet2=workbook['CR IN BANK'],
                                  sheet3=workbook['DR IN LEDGER'],
                                  sheet4=workbook['CR IN LEDGER'],
                                  destination_sheet=workbook['Knock off data'],
                                  top_sheet_date=recon_date
                                  )

                workbook.save(file_path)

                print('CR IN LEDGER (PENNANT FILE) - DONE')
        
        except Exception as e:
            print(f"str{e}")
    #==================================main processing ended=============================#

        try:
            #Create processed file object
            processed_files = ProcessedFiles(
                description=f"{os.path.basename(file_path)}",
                url=file_path)

            #add to the 
            db.session.add(processed_files)
            db.session.commit()
        except Exception as e:
            print(f"str{e}")

        end = time.time()
        total = end - start 
        print('took', total, 'seconds')
        flash(f"Processing Completed in {total} seconds. Your file is ready to download", 'success')
    else:
        flash("Select files", 'danger')
    return redirect(url_for('home'))
    
#================================Download and Delete files routes===================#    
       
#download link through flask app
@app.route('/download/<path:filepath>')
def download_file(filepath):
    try:
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        flash(f"str{e}", 'danger')
        print(f"str{e}")
        return redirect(url_for('home'))  

#Delete file using file id
@app.route("/<int:id>/delete/", methods=['POST'])
def delete_processed_files(id):
    try:
        file = ProcessedFiles.query.get(id)
        db.session.delete(file)
        db.session.commit()
        flash('file removed', 'success')
        return redirect(url_for('home'))
    except Exception as e:
        print(f"str{e}")
        flash(f'str{e}', 'danger')
        return redirect(url_for('home'))

#Delete all files in one shot
@app.route("/delete_all/", methods=['POST'])
def delete_all_files():
    try:
        db.session.query(ProcessedFiles).delete()
        db.session.commit()
        flash('All files removed', 'success')
        return redirect(url_for('home'))
    except Exception as e:
        print(f"str{e}")
        flash(f'str{e}', 'danger')
        return redirect(url_for('home'))

#================================Download and Delete files routes===================#

# to save file and return dictionary
def save_file(form_file):
    random_hex = secrets.token_hex(4)
    _, f_name = os.path.split(form_file.filename)
    file_name, file_ext = os.path.splitext(
        f_name)
    file_name = f"{file_name}_{random_hex}{file_ext}"
    file_path = os.path.join(app.root_path, 'media/uploaded_files', file_name)
    form_file.save(file_path)
    real_name = file_name
    return {'name': real_name, 'path': file_path}
