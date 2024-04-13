import os
from flask_app import app
from openpyxl import load_workbook

#load master workbook
def load_mastersheet(file):
    try:
        # target_workbook_path = os.path.join(app.root_path, 'media', 'output file', 'BRS -14-FEB-2024 - (HDFCCOLL4310) -BFL-HDFC-UPI COLLECTION INTEGRATION 4310.xlsx')
        workbook = load_workbook(file)
        return {'workbook': workbook}
    except Exception as e:
        print(f"str{e}")

#update recon date in top sheet D2 cell
def update_recon_date(workbook, recon_date):
    try:
        mastersheet = workbook.get('workbook')
        mastersheet['TOP SHEET']['D2'] = recon_date
    except Exception as e:
        print(f"str{e}")

#load input file
def load_input_file(file):
    try:
        source_workbook = load_workbook(file, data_only=True)
        source_ws = source_workbook.active
        return source_ws
    except Exception as e:
        print(f"str{e}")

